"""Ingestion + cleaning for the problem-statement corpus.

Implements the Section 3 rules of the build brief:

* Header row is discovered, not assumed (the source workbook has a 7-row title block).
* Duplicate ProjectID  -> keep both rows, suffix the later internal id (`-B`, `-C` ...),
  flag both as ``duplicate_project_id``.
* Missing / blank Title -> flag ``missing_title`` (do not drop).
* Missing / non-text Problem Statement (incl. literal numbers, or a statement that
  only echoes the title) -> flag ``missing_statement`` (do not drop).
* Flagged rows stay in the database but are excluded from the active
  recommendation pool; they surface in the admin "Needs Review" queue.

The same code path handles the initial workbook and later uploaded batches
(xlsx or csv) so retraining is a repeatable pipeline, not a one-off script.
"""

from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

# Minimum characters for a Problem Statement to count as "real text".
MIN_STATEMENT_CHARS = 25
# A title needs at least one word of this many letters to be meaningful
# (the source has rows titled "v" and "uu" with a real statement attached).
MIN_TITLE_WORD_LETTERS = 3

_HEADER_ALIASES = {
    "project_id": {"projectid", "project id", "project_id", "id", "pid"},
    "title": {"title", "problem title", "project title"},
    "statement": {"problem statement", "statement", "problem_statement", "description"},
}


@dataclass
class CleanedRow:
    project_id: str
    title: str | None
    statement: str | None
    is_flagged: bool
    flag_reason: str | None
    original_project_id: str | None
    source_row: int  # 1-based row number in the source sheet, for the review queue


@dataclass
class IngestReport:
    source: str
    sheet: str | None
    total_rows: int = 0
    active_rows: int = 0
    flagged_rows: int = 0
    flag_counts: dict[str, int] = field(default_factory=dict)
    flagged_detail: list[dict[str, Any]] = field(default_factory=list)

    def add_flag(self, reason: str) -> None:
        for part in reason.split("; "):
            self.flag_counts[part] = self.flag_counts.get(part, 0) + 1

    def as_dict(self) -> dict[str, Any]:
        return {
            "source": self.source,
            "sheet": self.sheet,
            "total_rows": self.total_rows,
            "active_rows": self.active_rows,
            "flagged_rows": self.flagged_rows,
            "flag_counts": self.flag_counts,
            "flagged_detail": self.flagged_detail,
        }


def _norm_header(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def _match_columns(header: list[Any]) -> dict[str, int] | None:
    normed = [_norm_header(h) for h in header]
    mapping: dict[str, int] = {}
    for key, aliases in _HEADER_ALIASES.items():
        for idx, name in enumerate(normed):
            if name in aliases:
                mapping[key] = idx
                break
    if {"project_id", "title", "statement"} <= mapping.keys():
        return mapping
    return None


def _clean_text(value: Any) -> str | None:
    """Return trimmed text, or ``None`` for blanks / non-text / numeric junk cells."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return None
    text = str(value).replace("\xa0", " ").strip()
    if not text or text.lower() in {"nan", "none", "n/a", "na", "-", "null"}:
        return None
    return text


def _collapse(text: str) -> str:
    return re.sub(r"\s+", " ", text).strip()


def _is_meaningful_title(title: str | None) -> bool:
    if not title:
        return False
    return any(
        len(re.sub(r"[^A-Za-z]", "", word)) >= MIN_TITLE_WORD_LETTERS
        for word in title.split()
    )


# --------------------------------------------------------------------------- #
# raw readers
# --------------------------------------------------------------------------- #
def _read_xlsx(data: bytes, preferred_sheet: str | None) -> tuple[str, list[list[Any]]]:
    wb = load_workbook(io.BytesIO(data), data_only=True)
    names = wb.sheetnames
    ordered = ([preferred_sheet] if preferred_sheet in names else []) + [
        n for n in names if n != preferred_sheet
    ]
    for name in ordered:
        rows = [list(r) for r in wb[name].iter_rows(values_only=True)]
        for header_idx in range(min(15, len(rows))):
            mapping = _match_columns(rows[header_idx])
            if mapping:
                body = rows[header_idx + 1 :]
                return name, _project(body, mapping, start_row=header_idx + 2)
    raise ValueError(
        "Could not find a header row with ProjectID / Title / Problem Statement "
        f"columns in any sheet of the workbook (sheets: {names})."
    )


def _read_csv(data: bytes) -> tuple[str, list[list[Any]]]:
    text = data.decode("utf-8-sig", errors="replace")
    reader = list(csv.reader(io.StringIO(text)))
    for header_idx in range(min(15, len(reader))):
        mapping = _match_columns(reader[header_idx])
        if mapping:
            return "csv", _project(reader[header_idx + 1 :], mapping, start_row=header_idx + 2)
    raise ValueError("Could not find ProjectID / Title / Problem Statement columns in the CSV.")


def _project(
    body: list[list[Any]], mapping: dict[str, int], start_row: int
) -> list[list[Any]]:
    out: list[list[Any]] = []
    for offset, row in enumerate(body):
        def cell(key: str) -> Any:
            idx = mapping[key]
            return row[idx] if idx < len(row) else None

        pid, title, stmt = cell("project_id"), cell("title"), cell("statement")
        if pid is None and title is None and stmt is None:
            continue  # trailing blank row
        out.append([pid, title, stmt, start_row + offset])
    return out


# --------------------------------------------------------------------------- #
# public API
# --------------------------------------------------------------------------- #
def read_source(path: str | Path, preferred_sheet: str | None = "QUANTUM PROBLEM TITLS & STATMNT"):
    path = Path(path)
    data = path.read_bytes()
    return _dispatch(data, path.name, preferred_sheet)


def read_upload(data: bytes, filename: str, preferred_sheet: str | None = None):
    return _dispatch(data, filename, preferred_sheet)


def _dispatch(data: bytes, filename: str, preferred_sheet: str | None):
    lower = filename.lower()
    if lower.endswith((".xlsx", ".xlsm", ".xls")):
        return _read_xlsx(data, preferred_sheet)
    if lower.endswith(".csv"):
        return _read_csv(data)
    raise ValueError(f"Unsupported file type: {filename!r} (expected .xlsx or .csv)")


def clean(
    raw_rows: list[list[Any]],
    *,
    source: str,
    sheet: str | None,
    source_batch: str,
) -> tuple[list[CleanedRow], IngestReport]:
    """Apply the Section 3 cleaning rules. Returns cleaned rows + a report."""
    report = IngestReport(source=source, sheet=sheet)
    cleaned: list[CleanedRow] = []
    seen_ids: dict[str, list[int]] = {}  # base project_id -> indices into `cleaned`

    for pid_raw, title_raw, stmt_raw, source_row in raw_rows:
        report.total_rows += 1
        base_pid = _clean_text(pid_raw) or f"UNKNOWN-ROW-{source_row}"
        title = _clean_text(title_raw)
        statement = _clean_text(stmt_raw)

        reasons: list[str] = []
        original_pid: str | None = None
        project_id = base_pid

        # --- duplicate ProjectID ------------------------------------------- #
        if base_pid in seen_ids:
            suffix = chr(ord("B") + len(seen_ids[base_pid]) - 1)  # 2nd -> B, 3rd -> C ...
            project_id = f"{base_pid}-{suffix}"
            original_pid = base_pid
            reasons.append("duplicate_project_id")
            # retro-flag the first occurrence
            first = cleaned[seen_ids[base_pid][0]]
            if "duplicate_project_id" not in (first.flag_reason or ""):
                first.flag_reason = _join(first.flag_reason, "duplicate_project_id")
                first.is_flagged = True
                report.add_flag("duplicate_project_id")
                _sync_detail(report, first)
            seen_ids[base_pid].append(len(cleaned))
        else:
            seen_ids[base_pid] = [len(cleaned)]

        # --- missing title ------------------------------------------------- #
        if not _is_meaningful_title(title):
            reasons.append("missing_title")

        # --- missing / non-text statement -------------------------------- #
        if statement is None or len(statement) < MIN_STATEMENT_CHARS:
            reasons.append("missing_statement")
        elif title is not None and _collapse(statement).lower() == _collapse(title).lower():
            # statement only echoes the title - no real problem description
            reasons.append("missing_statement")

        flag_reason = "; ".join(dict.fromkeys(reasons)) or None
        row = CleanedRow(
            project_id=project_id,
            title=title,
            statement=statement,
            is_flagged=flag_reason is not None,
            flag_reason=flag_reason,
            original_project_id=original_pid,
            source_row=source_row,
        )
        cleaned.append(row)

        if flag_reason:
            for part in flag_reason.split("; "):
                report.add_flag(part)
            _sync_detail(report, row)

    report.flagged_rows = sum(1 for r in cleaned if r.is_flagged)
    report.active_rows = len(cleaned) - report.flagged_rows
    return cleaned, report


def _join(existing: str | None, new: str) -> str:
    parts = [p for p in (existing or "").split("; ") if p]
    if new not in parts:
        parts.append(new)
    return "; ".join(parts)


def _sync_detail(report: IngestReport, row: CleanedRow) -> None:
    entry = {
        "source_row": row.source_row,
        "project_id": row.project_id,
        "title": row.title,
        "flag_reason": row.flag_reason,
    }
    for i, d in enumerate(report.flagged_detail):
        if d["source_row"] == row.source_row:
            report.flagged_detail[i] = entry
            return
    report.flagged_detail.append(entry)


def ingest_file(
    data: bytes, filename: str, *, source_batch: str, preferred_sheet: str | None = None
) -> tuple[list[CleanedRow], IngestReport]:
    sheet, raw = read_upload(data, filename, preferred_sheet)
    return clean(raw, source=filename, sheet=sheet, source_batch=source_batch)
